"""
server.py — TikTok Audio Dashboard V2.1 (All-in-One)
=====================================================
Chạy: python server.py
→ Khởi động cả 3 engine:
  - Creator Mining (Profile Engine) + Keyword Explorer (tool_nhac)
  - FYP Scroll (tool_sroll_feed) → subprocess
  - Background Re-check Job → auto mỗi 6 tiếng
→ Mở http://localhost:8000 để điều khiển.
"""
import asyncio
import json
import sys
import re
import subprocess
import aiosqlite
import warnings
import logging
from pathlib import Path
from typing import Optional, Callable
from contextlib import asynccontextmanager

# Chặn các cảnh báo rác của asyncio trên Windows khi tắt subprocess
warnings.filterwarnings("ignore", category=ResourceWarning)
# Chặn log lỗi 'I/O operation on closed pipe' gây rối mắt khi tắt server
logging.getLogger("asyncio").setLevel(logging.CRITICAL)


import uvicorn
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException, Request
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

# ══════════════════════════════════════════════════════════════════════════════
# PATHS
# ══════════════════════════════════════════════════════════════════════════════
BASE_DIR      = Path(__file__).parent
FYP_DIR       = BASE_DIR.parent / "tool_sroll_feed"
FYP_MAIN      = FYP_DIR / "main.py"
TOOL_NHAC_DB  = BASE_DIR / "data" / "database" / "audio_automation.db"
TOOL_FYP_DB   = FYP_DIR / "tiktok_audio.db"

# Đảm bảo tool_nhac được tìm thấy khi import
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

# Import run_crawler ở module level — tránh lazy import fail silently
try:
    from main import run_crawler as _run_crawler
    print(f"✅ Đã import run_crawler từ {BASE_DIR / 'main.py'}")
except Exception as _import_err:
    print(f"❌ Không import được run_crawler: {_import_err}")
    _run_crawler = None

# ══════════════════════════════════════════════════════════════════════════════
# GLOBAL STATE
# ══════════════════════════════════════════════════════════════════════════════
_connected:           set   = set()
_results:             list  = []
_nhac_log:            list  = []   # log của tool_nhac
_fyp_log:             list  = []   # log của tool_sroll_feed
_miner_log:           list  = []   # log của creator_miner
_follow_log:          list  = []   # log của auto_follow
_expander_log:        list  = []   # log của channel_expander
_status:              str   = "idle"
_session_count:       int   = 0
_target:              int   = 9999
_niche:               str   = "auto"
_nhac_task:           Optional[asyncio.Task] = None
_stop_event:          Optional[asyncio.Event] = None
_fyp_proc:            Optional[subprocess.Popen] = None
_fyp_reader_task:     Optional[asyncio.Task] = None
_follow_reader_task:  Optional[asyncio.Task] = None
_expander_task:       Optional[asyncio.Task] = None
_recheck_task:        Optional[asyncio.Task] = None
_recheck_interval:    int   = 360  # phút — chạy background recheck mỗi 6 tiếng

# Chuông báo cho Phase 3 (V2.2 Smart Pipeline)
_ai_wakeup = asyncio.Event()

MAX_LOG = 500

# ══════════════════════════════════════════════════════════════════════════════
# BROADCAST
# ══════════════════════════════════════════════════════════════════════════════
async def _broadcast(msg: dict):
    dead = set()
    for ws in _connected:
        try:
            await ws.send_text(json.dumps(msg, ensure_ascii=False))
        except Exception:
            dead.add(ws)
    _connected.difference_update(dead)

def _emit_log(msg: str, source: str = "nhac"):
    """Ghi log vào buffer và broadcast lên tất cả WebSocket clients."""
    print(f"[{source.upper()}] {msg}")
    
    if source == "nhac":     buf = _nhac_log
    elif source == "fyp":    buf = _fyp_log
    elif source == "miner":  buf = _miner_log
    elif source == "follow": buf = _follow_log
    else:                    buf = _expander_log
    buf.append(msg)
    if len(buf) > MAX_LOG:
        buf.pop(0)
    try:
        asyncio.get_event_loop().create_task(
            _broadcast({"type": "log", "source": source, "msg": msg})
        )
    except RuntimeError:
        pass

def _on_nhac_log(data):
    if isinstance(data, str):
        _emit_log(data, "nhac")
    else:
        asyncio.get_event_loop().create_task(_broadcast(data))

def _on_nhac_result(row):
    global _results, _session_count
    audio_id = row.get("audio_id", "") if isinstance(row, dict) else getattr(row, "audio_id", "")
    entry = {
        "audio_id":       audio_id,
        "audio_name":     (row.get("audio_name") if isinstance(row, dict) else getattr(row, "audio_name", "")) or "Unknown",
        "usage_count":    int((row.get("usage_count") if isinstance(row, dict) else getattr(row, "usage_count", 0)) or 0),
        "audio_page_url": (row.get("audio_page_url") if isinstance(row, dict) else getattr(row, "audio_page_url", "")) or "",
        "video_url":      (row.get("video_url") if isinstance(row, dict) else getattr(row, "video_url", "")) or "",
        "ai_score":       round(float((row.get("ai_score") if isinstance(row, dict) else getattr(row, "ai_score", 0)) or 0), 2),
        "speech_ratio":   round(float((row.get("speech_ratio") if isinstance(row, dict) else getattr(row, "speech_ratio", 0)) or 0) * 100),
        "source":         (row.get("source_type") if isinstance(row, dict) else getattr(row, "source_type", "keyword")) or "keyword",
        "status":         (row.get("status") if isinstance(row, dict) else getattr(row, "status", "accepted")) or "accepted",
        "date_added":     (row.get("date_added") if isinstance(row, dict) else getattr(row, "date_added", "")) or "",
        "origin":         "tool_nhac",
    }
    if not any(r["audio_id"] == audio_id for r in _results):
        _results.insert(0, entry)
        _results.sort(key=lambda x: x["usage_count"], reverse=True)
    _session_count += 1
    asyncio.get_event_loop().create_task(
        _broadcast({"type": "result", "data": entry, "total": _session_count, "target": _target})
    )

# ══════════════════════════════════════════════════════════════════════════════
# TOOL_NHAC RUNNER (chạy trong cùng event loop - không giới hạn target)
# ══════════════════════════════════════════════════════════════════════════════
async def _nhac_infinite_loop(niche: str = "auto"):
    """Chạy tool_nhac liên tục, tự khởi động lại nếu crash."""
    global _status
    while True:
        if _stop_event and _stop_event.is_set():
            _status = "stopped"
            _emit_log("⏹ Global Seeder đã dừng.", "nhac")
            await _broadcast({"type": "status", "nhac_status": "stopped", "fyp_status": _fyp_status_str()})
            return
        try:
            _status = "running"
            _emit_log(f"🚀 Global Seeder khởi động (Ngách: {niche})...", "nhac")
            await _broadcast({"type": "status", "nhac_status": "running", "fyp_status": _fyp_status_str()})
            if _run_crawler is None:
                raise RuntimeError("Không import được run_crawler từ main.py")
            await _run_crawler(
                target=_target,
                niche=niche,
                on_log=_on_nhac_log,
                on_result=_on_nhac_result,
                stop_event=_stop_event,
            )
            # Nếu hết target → reset và chạy lại
            _status = "done"
            await _broadcast({"type": "status", "nhac_status": "done", "fyp_status": _fyp_status_str()})
            _emit_log("✅ Hoàn tất 1 vòng, tự động bắt đầu vòng mới sau 10s...", "nhac")
            await asyncio.sleep(10)
        except asyncio.CancelledError:
            _status = "stopped"
            await _broadcast({"type": "status", "nhac_status": "stopped", "fyp_status": _fyp_status_str()})
            return
        except Exception as e:
            _status = "error"
            err_msg = f"💥 Lỗi tool_nhac: {e} — Tự restart sau 30s..."
            print(err_msg)  # In ra terminal để debug
            _emit_log(err_msg, "nhac")
            await _broadcast({"type": "status", "nhac_status": "error", "fyp_status": _fyp_status_str()})
            await asyncio.sleep(30)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL_SROLL_FEED RUNNER (subprocess)
# ══════════════════════════════════════════════════════════════════════════════
def _fyp_running() -> bool:
    return _fyp_proc is not None and _fyp_proc.poll() is None

def _fyp_status_str() -> str:
    return "running" if _fyp_running() else "stopped"

def _start_fyp_proc():
    global _fyp_proc
    if _fyp_running():
        return False, "FYP đang chạy rồi"
    if not FYP_MAIN.exists():
        return False, f"Không tìm thấy {FYP_MAIN}"
    _fyp_proc = subprocess.Popen(
        [sys.executable, "-u", str(FYP_MAIN)],
        cwd=str(FYP_DIR),
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        text=True, bufsize=1, encoding="utf-8", errors="replace"
    )
    return True, "FYP tool đã khởi động"

def _stop_fyp_proc():
    global _fyp_proc
    if _fyp_proc and _fyp_proc.poll() is None:
        try:
            import platform
            if platform.system() == "Windows":
                subprocess.call(['taskkill', '/F', '/T', '/PID', str(_fyp_proc.pid)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:
                _fyp_proc.kill()
        except Exception:
            try: _fyp_proc.kill()
            except: pass
    _fyp_proc = None
    return True, "FYP tool đã dừng"

_follow_proc = None

def _follow_running() -> bool:
    return _follow_proc is not None and _follow_proc.poll() is None

def _start_follow_proc(target: str, max_follows: int):
    global _follow_proc
    if _follow_running():
        return False, "Auto Follow đang chạy rồi"
    script_path = FYP_DIR / "auto_follower.py"
    if not script_path.exists():
        return False, f"Không tìm thấy {script_path}"
    _follow_proc = subprocess.Popen(
        [sys.executable, "-u", str(script_path), "--target", target, "--max", str(max_follows)],
        cwd=str(FYP_DIR),
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        text=True, bufsize=1, encoding="utf-8", errors="replace"
    )
    return True, "Auto Follow đã khởi động"

def _stop_follow_proc():
    global _follow_proc
    if _follow_proc and _follow_proc.poll() is None:
        try:
            import platform
            if platform.system() == "Windows":
                subprocess.call(['taskkill', '/F', '/T', '/PID', str(_follow_proc.pid)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:
                _follow_proc.kill()
        except Exception:
            try: _follow_proc.kill()
            except: pass
    _follow_proc = None
    return True, "Auto Follow đã dừng"

async def _fyp_log_streamer():
    """
    Chỉ stream stdout của FYP process về dashboard.
    Không tự restart — người dùng toàn quyền điều khiển qua web.
    """
    while True:
        await asyncio.sleep(0.3)
        if _fyp_proc and _fyp_proc.stdout:
            try:
                line = _fyp_proc.stdout.readline()
                if line:
                    line = line.strip()
                    if line.startswith("RESULT_JSON:"):
                        try:
                            data = json.loads(line[12:])
                            entry = {
                                "audio_id":       data.get("audio_id", ""),
                                "audio_name":     data.get("audio_name", "FYP Audio"),
                                "usage_count":    data.get("usage_count", 0),
                                "audio_page_url": data.get("audio_link", ""),
                                "video_url":      data.get("video_link", ""),
                                "ai_score":       0,
                                "speech_ratio":   0,
                                "source":         "fyp",
                                "date_added":     data.get("created_at", ""),
                                "origin":         "tool_sroll"
                            }
                            global _results, _session_count
                            if not any(r["audio_id"] == entry["audio_id"] for r in _results):
                                _results.insert(0, entry)
                                _results.sort(key=lambda x: x["usage_count"], reverse=True)
                            
                            # Broadcast result to update UI
                            await _broadcast({"type": "result", "data": entry, "total": _session_count, "target": _target})
                        except Exception as e:
                            print(f"Lỗi parse RESULT_JSON từ FYP: {e}")
                    else:
                        _fyp_log.append(line)
                        if len(_fyp_log) > MAX_LOG: _fyp_log.pop(0)
                        await _broadcast({"type": "log", "source": "fyp", "msg": line})
            except Exception:
                pass

async def _follow_log_streamer():
    """Stream stdout của Auto Follower chung vào fyp log window."""
    while True:
        await asyncio.sleep(0.3)
        if _follow_proc and _follow_proc.stdout:
            try:
                line = _follow_proc.stdout.readline()
                if line:
                    line = line.strip()
                    _fyp_log.append("[FOLLOW] " + line)
                    if len(_fyp_log) > MAX_LOG: _fyp_log.pop(0)
                    await _broadcast({"type": "log", "source": "follow", "msg": "[FOLLOW] " + line})
            except Exception:
                pass

# ══════════════════════════════════════════════════════════════════════════════
# DB LOADER (đọc kết quả từ cả 2 DB)
# ══════════════════════════════════════════════════════════════════════════════
async def _load_all_results() -> list:
    results = []

    if TOOL_NHAC_DB.exists():
        try:
            async with aiosqlite.connect(str(TOOL_NHAC_DB), timeout=10) as db:
                db.row_factory = aiosqlite.Row
                async with db.execute(
                    "SELECT * FROM audio_history WHERE status IN ('accepted', 'pending_ai') ORDER BY date_added DESC LIMIT 2000"
                ) as cur:
                    async for row in cur:
                        keys = row.keys()
                        source = row["source_type"] if "source_type" in keys else "keyword"
                        if source == "creator_scan": source = "profile"
                        
                        results.append({
                            "audio_id":       row["audio_id"],
                            "audio_name":     row["audio_name"] or "Unknown",
                            "usage_count":    row["usage_count"] or 0,
                            "audio_page_url": row["audio_page_url"] or "",
                            "video_url":      row["video_url"] if "video_url" in keys else "",
                            "ai_score":       round(float(row["ai_score"] or 0), 2) if "ai_score" in keys else 0,
                            "speech_ratio":   round(float(row["speech_ratio"] or 0) * 100) if "speech_ratio" in keys else 0,
                            "source":         source,
                            "status":         row["status"] if "status" in keys else "accepted",

                            "date_added":     row["date_added"] if "date_added" in keys else "",
                            "origin":         "tool_nhac",
                        })
        except Exception as e:
            print(f"⚠️ Lỗi đọc tool_nhac DB: {e}")

    if TOOL_FYP_DB.exists():
        try:
            async with aiosqlite.connect(str(TOOL_FYP_DB), timeout=10) as db:
                db.row_factory = aiosqlite.Row
                async with db.execute(
                    "SELECT * FROM audiorecord WHERE status='passed' ORDER BY created_at DESC LIMIT 2000"
                ) as cur:
                    async for row in cur:
                        keys = row.keys()
                        results.append({
                            "audio_id":       row["audio_id"],
                            "audio_name":     row["audio_name"] if "audio_name" in keys else "FYP Audio",
                            "usage_count":    row["usage_count"] or 0,
                            "audio_page_url": row["audio_link"] if "audio_link" in keys else "",
                            "video_url":      row["original_video_link"] if "original_video_link" in keys else "",
                            "ai_score":       0,
                            "speech_ratio":   0,
                            "source":         "fyp",
                            "status":         "pending_ai",
                            "date_added":     str(row["created_at"]) if "created_at" in keys else "",
                            "origin":         "tool_sroll",
                        })
        except Exception as e:
            print(f"⚠️ Lỗi đọc FYP DB: {e}")

    seen, deduped = set(), []
    for r in sorted(results, key=lambda x: x["usage_count"], reverse=True):
        if r["audio_id"] not in seen:
            seen.add(r["audio_id"])
            deduped.append(r)
    return deduped

def _get_stats() -> dict:
    by_src: dict = {}
    fyp = nhac = 0
    for r in _results:
        s = r.get("source", "keyword")
        by_src[s] = by_src.get(s, 0) + 1
        if r["origin"] == "tool_sroll": fyp += 1
        else: nhac += 1
    # V2.1: thêm source mới
    return {
        "total": len(_results), "fyp": fyp, "nhac": nhac,
        "by_source": by_src,
        "profile": by_src.get("profile", 0),
        "hot_trend": sum(1 for r in _results if r.get("trend_tag") in ("HOT_TREND", "EARLY_TREND")),
    }


def _creators_count() -> int:
    """Đếm số creator trong creators_list.txt."""
    try:
        p = BASE_DIR / "creators_list.txt"
        if p.exists():
            return sum(1 for l in p.read_text(encoding='utf-8').splitlines() if l.strip() and not l.startswith('#'))
    except Exception:
        pass
    return 0

# ══════════════════════════════════════════════════════════════════════════════
# LIFESPAN — Tất cả khởi động ở đây
# ══════════════════════════════════════════════════════════════════════════════
@asynccontextmanager
async def lifespan(app: FastAPI):
    global _results, _nhac_task, _stop_event, _fyp_reader_task

    print("\n" + "═"*60)
    print("  🎵 TikTok Audio Dashboard — All-in-One")
    print("  http://localhost:8000")
    print("═"*60)

    # Nạp kết quả từ cả 2 DB
    _results = await _load_all_results()
    stats = _get_stats()
    print(f"  📊 Đã nạp {stats['total']} kết quả (tool_nhac: {stats['nhac']}, FYP: {stats['fyp']}, Profile: {stats['profile']})")
    print(f"  📋 Danh sách creator: {_creators_count()} kênh")

    # Khởi động FYP log streamer
    _fyp_reader_task = asyncio.get_event_loop().create_task(_fyp_log_streamer())
    _follow_reader_task = asyncio.get_event_loop().create_task(_follow_log_streamer())
    _stop_event = asyncio.Event()

    # V2.1: Tự động chạy Background Recheck mỗi 6 tiếng
    # _recheck_task = asyncio.get_event_loop().create_task(_auto_recheck_loop())
    _recheck_task = None

    # Khởi động AI Check Worker (Phase 3)
    global _ai_worker_task
    _ai_worker_task = asyncio.get_event_loop().create_task(_ai_check_worker())

    print("  ✅ Dashboard sẵn sàng — Bấm nút trên web để bắt đầu các tool")
    print("═"*60 + "\n")

    yield  # Server đang chạy

    # Cleanup khi tắt
    print("\n🛑 Đang tắt hệ thống...")
    if _stop_event: _stop_event.set()
    
    # Cố gắng đóng các task êm ái
    tasks = [t for t in [_nhac_task, _fyp_reader_task, _follow_reader_task, _expander_task, _recheck_task, _ai_worker_task] if t and not t.done()]
    for t in tasks: t.cancel()
    
    _stop_fyp_proc()
    
    # Một chút delay để pipe kịp đóng trước khi loop bị hủy
    await asyncio.sleep(0.5)
    print("✅ Đã dừng tất cả.")


# ══════════════════════════════════════════════════════════════════════════════
# FASTAPI APP
# ══════════════════════════════════════════════════════════════════════════════
app = FastAPI(title="TikTok Audio Dashboard", lifespan=lifespan)

# ── REST: tool_nhac control ────────────────────────────────────────────────────
@app.post("/start")
async def nhac_start(body: dict = {}):
    global _nhac_task, _stop_event, _status, _session_count, _target, _niche
    if _nhac_task and not _nhac_task.done():
        return {"ok": False, "msg": "Global Seeder đang chạy rồi"}
    _target = int(body.get("target", 20))
    _niche = body.get("niche", "auto")
    _session_count = 0
    _stop_event = asyncio.Event()
    _nhac_task = asyncio.get_event_loop().create_task(_nhac_infinite_loop(_niche))
    # Broadcast ngay lập tức — không chờ WS polling
    await _broadcast({"type": "status", "nhac_status": "running", "fyp_status": _fyp_status_str()})
    return {"ok": True, "msg": f"Global Seeder đã khởi động (target={_target}, niche={_niche})"}

@app.post("/stop")
async def nhac_stop():
    global _status, _nhac_task
    if _stop_event: _stop_event.set()
    if _nhac_task and not _nhac_task.done():
        _nhac_task.cancel()
    _status = "stopped"
    await _broadcast({"type": "status", "nhac_status": "stopped", "fyp_status": _fyp_status_str()})
    return {"ok": True, "msg": "tool_nhac đã dừng"}

# ── REST: FYP tool control ─────────────────────────────────────────────────────
@app.post("/fyp/start")
async def fyp_start_api():
    ok, msg = _start_fyp_proc() if not _fyp_running() else (False, "Đang chạy rồi")
    await _broadcast({"type": "fyp_status", "status": _fyp_status_str(), "msg": msg})
    return {"ok": ok, "msg": msg}

@app.post("/fyp/stop")
async def fyp_stop_api():
    # Tạm thời dừng FYP (watchdog sẽ không restart khi stop_event được set)
    _stop_fyp_proc()
    await _broadcast({"type": "fyp_status", "status": "stopped", "msg": "FYP đã dừng"})
    return {"ok": True, "msg": "FYP tool đã dừng"}

@app.get("/fyp/status")
async def fyp_status_api():
    return {"status": _fyp_status_str(), "running": _fyp_running()}

@app.post("/fyp/result")
async def fyp_result_api(body: dict):
    """
    Nhận kết quả từ tool_sroll_feed qua HTTP POST.
    Thay thế cơ chế đọc stdout (RESULT_JSON:...) — đáng tin cậy hơn.
    """
    global _results, _session_count
    try:
        audio_id = body.get("audio_id", "")
        if not audio_id:
            return {"ok": False, "msg": "Missing audio_id"}

        source = body.get("source", "fyp")
        if source == "creator_scan": source = "profile"

        entry = {
            "audio_id":       audio_id,
            "audio_name":     body.get("audio_name", "FYP Audio"),
            "usage_count":    int(body.get("usage_count", 0) or 0),
            "audio_page_url": body.get("audio_link", ""),
            "video_url":      body.get("video_link", ""),
            "ai_score":       0,
            "speech_ratio":   0,
            "source":         source,
            "status":         "pending_ai",
            "date_added":     body.get("created_at", ""),
            "origin":         "tool_sroll",
        }


        if not any(r["audio_id"] == audio_id for r in _results):
            _results.insert(0, entry)
            _results.sort(key=lambda x: x["usage_count"], reverse=True)

        _session_count += 1
        await _broadcast({"type": "result", "data": entry, "total": _session_count, "target": _target})
        
        # [V2.2 Smart] Đánh thức Phase 3 dậy làm việc ngay
        _ai_wakeup.set()
        
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}

# ── REST: Data & Status ────────────────────────────────────────────────────────
@app.get("/status")
async def get_status():
    return {
        "status":      _status,
        "nhac_status": "running" if (_nhac_task and not _nhac_task.done()) else "stopped",
        "fyp_status":  _fyp_status_str(),
        "total":       len(_results),
        "target":      _target,
        "stats":       _get_stats(),
    }

@app.get("/results")
async def get_results(source: str = "all", origin: str = "all", limit: int = 500):
    filtered = _results
    if source != "all": filtered = [r for r in filtered if r.get("source") == source]
    if origin != "all": filtered = [r for r in filtered if r.get("origin") == origin]
    return {"results": filtered[:limit], "total": len(filtered), "stats": _get_stats()}

@app.get("/refresh")
async def refresh_db():
    global _results
    _results = await _load_all_results()
    stats = _get_stats()
    await _broadcast({"type": "refresh", "total": len(_results), "stats": stats,
                      "fyp_status": _fyp_status_str(), "nhac_status": "running" if (_nhac_task and not _nhac_task.done()) else "stopped"})
    return {"ok": True, "total": len(_results), "stats": stats}

@app.get("/get-expanded-creators")
async def get_expanded_creators():
    """Đọc file creators_list.txt và trả về danh sách username"""
    try:
        path = Path("creators_list.txt")
        if not path.exists():
            return {"ok": False, "msg": "File creators_list.txt không tồn tại (chưa cào kênh nào)."}
        
        content = path.read_text(encoding="utf-8").strip()
        if not content:
            return {"ok": False, "msg": "File creators_list.txt đang trống."}
            
        return {"ok": True, "data": content}
    except Exception as e:
        return {"ok": False, "msg": f"Lỗi đọc file: {e}"}

@app.delete("/delete/{audio_id}")
async def delete_audio_api(audio_id: str):
    global _results
    try:
        from database import delete_audio
        await delete_audio(audio_id)
        _results = [r for r in _results if r.get("audio_id") != audio_id]
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/expand-keywords")
async def expand_keywords():
    try:
        from ai.multi_ai_manager import MultiAIManager
        from config import KEYWORDS_FILE
        ai = MultiAIManager()
        if not ai.gemini.api_key and not ai.groq.api_key:
            return {"ok": False, "msg": "Chưa cấu hình AI API Key"}
        base = ai.get_random_keyword_from_file(KEYWORDS_FILE)
        keywords = await ai.expand_keyword(base)
        return {"ok": True, "base_keyword": base, "new_keywords": keywords}
    except Exception as e:
        return {"ok": False, "msg": str(e)}

@app.post("/audio/check-single")
async def audio_check_single(body: dict):
    """
    Kiểm tra 1 audio duy nhất từ URL (link video hoặc link nhạc).
    Luồng: Trích xuất music_id -> Check LSD (Phase 2) -> Check AI (Phase 3).
    """
    url = body.get("url", "").strip()
    if not url: return {"ok": False, "msg": "Vui lòng nhập URL TikTok"}
    
    _emit_log(f"🔎 Đang kiểm tra đơn lẻ: {url}", "checker")
    
    async def _run_check():
        from playwright.async_api import async_playwright
        from crawler import TikTokCrawler
        from models import AudioMetadata
        from audio_pipeline import AudioPipeline
        import urllib.parse

        try:
            from config import HEADLESS
            async with async_playwright() as pw:
                browser = await pw.chromium.launch(headless=HEADLESS, args=["--mute-audio"])
                context = await browser.new_context()
                page = await context.new_page()

                music_url = url
                audio_id = ""
                audio_name = "Single Check"

                # 1. Nếu là link video, trích xuất music_url
                if "/music/" in url:
                    _emit_log("🎵 Phát hiện link trang nhạc trực tiếp.", "checker")
                elif "/video/" in url:
                    _emit_log("🎬 URL là video, đang tìm link nhạc...", "checker")
                    await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                    # Chờ link nhạc xuất hiện (thử nhiều selector)
                    try:
                        music_link_el = None
                        selectors = ["a[href*='/music/']", "[data-e2e='video-music'] a", "[class*='StyledMusicLink']", "h4[data-e2e='music-title'] a", "a[href*='-music-']"]
                        for sel in selectors:
                            try:
                                music_link_el = await page.wait_for_selector(sel, timeout=5000)
                                if music_link_el: break
                            except: continue
                        
                        if music_link_el:
                            music_url = await music_link_el.get_attribute("href")
                        else:
                            # Thử tìm bằng regex trong toàn bộ HTML
                            html = await page.content()
                            match = re.search(r'\"music\":\{.*?\"url\":\"(.*?)\"', html)
                            if match:
                                music_url = match.group(1).replace("\\u002F", "/")
                            else:
                                match2 = re.search(r'/music/[\w-]+-\d+', html)
                                if match2: music_url = match2.group(0)

                        if music_url:
                            if music_url.startswith("/"): music_url = "https://www.tiktok.com" + music_url
                            _emit_log(f"🔗 Tìm thấy link nhạc: {music_url}", "checker")
                        else:
                            raise Exception("No music link found")
                    except Exception as e:
                        _emit_log(f"⚠️ Không tự động tìm được link nhạc: {e}. Thử dùng link hiện tại.", "checker")
                
                # 2. Trích xuất audio_id từ music_url hoặc url gốc
                audio_id = ""
                # Thử lấy ID từ music URL trước
                if "/music/" in music_url:
                    m = re.search(r"(\d+)(?:\?|$)", music_url)
                    if m: audio_id = m.group(1)
                
                # Nếu vẫn chưa có ID, thử lấy từ video URL
                if not audio_id:
                    m_vid = re.search(r"video/(\d+)", url)
                    if m_vid: 
                        audio_id = m_vid.group(1)
                        _emit_log(f"💡 Dùng ID video làm ID âm thanh: {audio_id}", "nhac")

                if not audio_id:
                    _emit_log("❌ Không trích xuất được ID hợp lệ từ URL này.", "nhac")
                    await browser.close()
                    return

                # 3. Phase 2: Check LSD (Usage Count)
                _emit_log(f"📊 [Phase 2] Đang check lượt dùng cho ID {audio_id}...", "nhac")
                crawler = TikTokCrawler()
                usage = await crawler.get_accurate_usage(music_url, page=page)
                _emit_log(f"📈 Lượt dùng: {usage:,}", "nhac")

                # 4. Trích xuất thêm thông tin
                _emit_log("🎵 Đang lấy thông tin âm thanh để chạy AI...", "nhac")
                try:
                    # Đánh chặn mạng để tìm link audio
                    captured_audio_url = ""
                    async def _handle_request(request):
                        nonlocal captured_audio_url
                        url_req = request.url
                        if ".mp3" in url_req or "audio" in request.resource_type or "mime_type=audio_mpeg" in url_req:
                            captured_audio_url = url_req
                    
                    page.on("request", _handle_request)

                    # Chờ thẻ audio xuất hiện hoặc có request audio
                    try:
                        await page.goto(music_url, wait_until="networkidle", timeout=30000)
                    except:
                        pass

                    # Cuộn nhẹ để trigger load audio player
                    await page.mouse.wheel(0, 500)
                    await asyncio.sleep(3)
                    
                    # Ưu tiên lấy từ network interception, nếu không có thì lấy từ DOM
                    audio_url = captured_audio_url
                    if not audio_url:
                        audio_url = await page.evaluate("() => document.querySelector('audio source')?.src || document.querySelector('audio')?.src || ''")
                    
                    if audio_url: 
                        _emit_log(f"✅ Đã lấy được link audio qua mạng", "nhac")
                    else:
                        _emit_log("⚠️ Cảnh báo: Không tìm thấy link tải nhạc (.mp3) qua cả mạng và DOM.", "nhac")
                    
                    # Thử lấy duration từ nhiều nguồn
                    duration = await page.evaluate("""() => {
                        const el = document.querySelector("[class*='MusicDuration'], [data-e2e='music-duration']");
                        if (el && el.innerText.includes(':')) {
                            const p = el.innerText.split(':');
                            return parseInt(p[0]) * 60 + parseInt(p[1]);
                        }
                        const a = document.querySelector('audio');
                        return a && a.duration ? Math.floor(a.duration) : 0;
                    }""")
                    
                    if duration <= 0: duration = 30 # Mặc định 30s để không bị chặn pipeline

                    real_name = await page.evaluate("() => document.querySelector('[data-e2e=\"music-title\"], h1')?.innerText || 'Single Check'")
                    if real_name: audio_name = real_name
                except Exception as e:
                    _emit_log(f"⚠️ Lỗi lấy metadata: {e}", "nhac")
                    audio_url = ""; duration = 30

                # 5. Tạo metadata
                audio = AudioMetadata(
                    audio_id=audio_id,
                    audio_name=audio_name,
                    duration=duration,
                    usage_count=usage,
                    audio_url=audio_url,
                    audio_page_url=music_url,
                    video_url=url if "/video/" in url else "",
                    status="pending_ai",
                    keyword="manual_check"
                )

                # 5.1 Re-validate (Bây giờ đã có audio_url, check xem có phải nhạc thư viện không)
                from filter import is_valid_audio
                ok, reason = is_valid_audio(audio)
                if not ok:
                    _emit_log(f"❌ KHÔNG ĐẠT: {reason}", "nhac")
                    await browser.close()
                    return

                # 6. Phase 3: AI Check
                _emit_log("🤖 [Phase 3] Đang chạy AI Check...", "nhac")
                pipeline = AudioPipeline.get()
                
                async def _on_acc(a):
                    _emit_log(f"✅ HỢP LỆ! (Score: {a.ai_score}, Speech: {round(a.speech_ratio*100)}%)", "nhac")
                    row = {
                        "audio_id": a.audio_id, "audio_name": a.audio_name, "usage_count": a.usage_count,
                        "ai_score": a.ai_score, "speech_ratio": a.speech_ratio, "status": "accepted",
                        "audio_page_url": a.audio_page_url, "source": "manual_check"
                    }
                    _results.insert(0, row)
                    asyncio.get_event_loop().create_task(_broadcast({"type": "result", "data": row}))

                result = await pipeline.process(audio, on_accepted=_on_acc)
                if not result.passed:
                    _emit_log(f"❌ KHÔNG ĐẠT: {result.reason}", "nhac")
                
                _emit_log("🏁 Kiểm tra đơn lẻ hoàn tất.", "nhac")

                await browser.close()

        except Exception as e:
            _emit_log(f"❌ Lỗi kiểm tra đơn lẻ: {e}", "nhac")

    asyncio.get_event_loop().create_task(_run_check())
    return {"ok": True, "msg": "Đang kiểm tra URL..."}

@app.post("/save-keywords")
async def save_keywords(body: dict):
    try:
        from config import KEYWORDS_FILE
        with open(KEYWORDS_FILE, "a", encoding="utf-8") as f:
            f.write("\n# --- AI EXPANDED ---\n")
            for kw in body.get("new_keywords", []):
                f.write(f"{kw}\n")
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}

@app.post("/follow/start")
async def start_follow(req: Request):
    data = await req.json()
    target = data.get("target", "")
    max_follows = int(data.get("max", 200))
    if not target:
        return {"ok": False, "msg": "Thiếu target username"}
    ok, msg = _start_follow_proc(target, max_follows)
    return {"ok": ok, "msg": msg}

@app.post("/follow/stop")
async def stop_follow():
    ok, msg = _stop_follow_proc()
    return {"ok": ok, "msg": msg}

# ── V2.1: Channel Expander API ─────────────────────────────────────────────────
async def _run_expander_task():
    global _expander_task
    _emit_log("🌱 Channel Expander đang chạy...", "expander")
    try:
        import sys
        from channel_expander import expand_creators
        await expand_creators()
        _emit_log(f"✅ Expander xong! {_creators_count()} kênh trong list.", "expander")
    except Exception as e:
        _emit_log(f"❌ Expander lỗi: {e}", "expander")
    finally:
        _expander_task = None

@app.post("/expander/start")
async def expander_start():
    global _expander_task
    if _expander_task and not _expander_task.done():
        return {"ok": False, "msg": "Expander đang chạy rồi"}
    _expander_task = asyncio.get_event_loop().create_task(_run_expander_task())
    return {"ok": True, "msg": "Channel Expander đã khởi động"}

@app.post("/expander/stop")
async def expander_stop():
    global _expander_task
    if _expander_task and not _expander_task.done():
        _expander_task.cancel()
        _expander_task = None
        _emit_log("🛑 Đã dừng Channel Expander.", "expander")
        return {"ok": True, "msg": "Đã dừng Channel Expander"}
    return {"ok": False, "msg": "Expander hiện không chạy"}

@app.get("/expander/status")
async def expander_status():
    running = _expander_task is not None and not _expander_task.done()
    return {"running": running, "creators_count": _creators_count()}


# ── V2.1: Background Recheck API ───────────────────────────────────────────────
async def _auto_recheck_loop():
    """Chạy background recheck tự động mỗi _recheck_interval phút."""
    await asyncio.sleep(60)  # Chờ 1 phút sau khi khởi động
    while True:
        _emit_log("🔄 [Auto] Background Re-check Job bắt đầu...", "nhac")
        try:
            from background_recheck import run_recheck
            await run_recheck()
            _emit_log("✅ [Auto] Re-check Job hoàn thành.", "nhac")
        except Exception as e:
            _emit_log(f"⚠️ [Auto] Re-check lỗi: {e}", "nhac")
        await asyncio.sleep(_recheck_interval * 60)

@app.post("/recheck/run")
async def recheck_now():
    """Chạy re-check thủ công ngay lập tức."""
    try:
        from background_recheck import run_recheck
        asyncio.get_event_loop().create_task(run_recheck())
        return {"ok": True, "msg": "Re-check đang chạy nền"}
    except Exception as e:
        return {"ok": False, "msg": str(e)}

# ── V2.1: PHASE 3 - AI Check Worker ───────────────────────────────────────────
_ai_worker_task: Optional[asyncio.Task] = None

async def _ai_check_worker():
    """
    Worker chạy ngầm, liên tục quét DB tìm audio 'pending_ai' để xử lý.
    Đây chính là 'Phase 3' mà người dùng mong đợi.
    """
    from database import get_pending_ai_audios, update_audio_ai_result
    from audio_pipeline import AudioPipeline
    
    pipeline = AudioPipeline.get()
    _emit_log("🤖 Phase 3 (AI Check Worker) đã sẵn sàng.", "nhac")
    
    while True:
        try:
            pending = await get_pending_ai_audios(limit=5)
            if not pending:
                # Log heartbeat mỗi 2 phút để xác nhận worker vẫn sống
                import time as _t
                if int(_t.time()) % 120 < 11:
                    _emit_log("🤖 Phase 3: Đang trực chiến (Đang đợi audio mới từ Phase 2...)", "nhac")
                
                # [V2.2 Smart] Thay vì sleep cố định 10s, ta đợi chuông báo hoặc tối đa 30s
                # await asyncio.sleep(10) # <-- Code cũ của bạn được comment lại ở đây
                try:
                    await asyncio.wait_for(_ai_wakeup.wait(), timeout=30.0)
                except asyncio.TimeoutError:
                    pass
                _ai_wakeup.clear() # Reset chuông sau khi dậy
                continue
            
            _emit_log(f"🤖 [Phase 3] Đang xử lý AI cho {len(pending)} audio...", "nhac")
            
            async def _proc(audio):
                try:
                    _emit_log(f"🔍 AI Check: {audio.audio_name[:30]}...", "nhac")
                    
                    def _on_accept(a):
                        _emit_log(f"✅ ĐẠT AI: {a.audio_name[:30]} (Speech: {round(a.speech_ratio*100)}%)", "nhac")
                        # Cập nhật kết quả lên UI ngay lập tức
                        row = {
                            "audio_id":     a.audio_id,
                            "audio_name":   a.audio_name,
                            "usage_count":  a.usage_count,
                            "ai_score":     a.ai_score,
                            "speech_ratio": a.speech_ratio,
                            "status":       "accepted",
                            "source":       a.source_type or "profile",
                        }
                        # Cập nhật vào list _results tại chỗ
                        for r in _results:
                            if r["audio_id"] == a.audio_id:
                                r.update(row)
                                break
                        
                        asyncio.get_event_loop().create_task(
                            _broadcast({"type": "result", "data": row})
                        )
                    
                    # Giới hạn 60s cho mỗi audio
                    pipeline_res = await asyncio.wait_for(
                        pipeline.process(audio, on_accepted=_on_accept),
                        timeout=60.0
                    )
                    
                    if not pipeline_res.passed:
                        _emit_log(f"❌ LOẠI AI: {audio.audio_name[:30]} | Lý do: {pipeline_res.reason}", "nhac")
                        # Cập nhật status rejected về UI để nó biến mất khỏi bảng "Chờ"
                        row = {"audio_id": audio.audio_id, "status": "rejected", "reason": pipeline_res.reason}
                        asyncio.get_event_loop().create_task(_broadcast({"type": "result", "data": row}))

                except asyncio.TimeoutError:
                    _emit_log(f"⚠️ AI Check quá lâu (60s), bỏ qua: {audio.audio_id}", "nhac")
                    audio.status = "rejected"
                    audio.reason = "AI Timeout"
                    await update_audio_ai_result(audio)
                except Exception as e:
                    _emit_log(f"❌ AI Check lỗi ({audio.audio_id}): {e}", "nhac")
                    audio.status = "rejected"
                    audio.reason = f"AI Error: {e}"
                    await update_audio_ai_result(audio)

            await asyncio.gather(*[_proc(a) for a in pending])
            # _emit_log(f"✅ [Phase 3] Xử lý xong lô {len(pending)} audio.", "nhac")


            
        except Exception as e:
            _emit_log(f"⚠️ [Phase 3] Lỗi worker: {e}", "nhac")
            await asyncio.sleep(30)
        
        await asyncio.sleep(2)


# ── V2.1: Stats & Trend endpoint ───────────────────────────────────────────────
@app.get("/v2/stats")
async def v2_stats():
    """Trả về stats đầy đủ V2.1 bao gồm trend tags và creator count."""
    stats = _get_stats()
    trend_tags = {}
    for r in _results:
        t = r.get("trend_tag", "NORMAL")
        trend_tags[t] = trend_tags.get(t, 0) + 1
    return {
        **stats,
        "trend_tags":     trend_tags,
        "creators_count": _creators_count(),
        "recheck_interval_min": _recheck_interval,
    }

@app.get("/creators")
async def get_creators():
    """Trả về danh sách creator trong creators_list.txt."""
    try:
        p = BASE_DIR / "creators_list.txt"
        if not p.exists():
            return {"creators": [], "total": 0}
        lines = [l.strip().lstrip('@') for l in p.read_text(encoding='utf-8').splitlines()
                 if l.strip() and not l.startswith('#')]
        return {"creators": lines, "total": len(lines)}
    except Exception as e:
        return {"error": str(e)}


# ── Creator Scanner API ────────────────────────────────────────────────────────
_scanner_task: Optional[asyncio.Task] = None
_scanner_log:  list = []

def _scanner_emit(msg: str):
    _scanner_log.append(msg)
    if len(_scanner_log) > 200:
        _scanner_log.pop(0)
    _emit_log(msg, "miner")  # Stream về log FYP trên dashboard

async def _run_creator_scan(usernames: list, force: bool = False):
    """Chạy CreatorScanner trong nền."""
    global _scanner_task
    try:
        import sys, traceback as tb
        sys.path.insert(0, str(BASE_DIR.parent / "tool_sroll_feed"))
        from creator_scanner import CreatorScanner, report_scan_result

        scanner = CreatorScanner(
            page         = None,
            checked_audio= set(r["audio_id"] for r in _results),
            on_result    = report_scan_result,
            log_fn       = _scanner_emit,
        )
        
        # [V2.2 Smart] Chạy cuốn chiếu từng user để Phase 3 có thể làm việc ngay
        # results = await scanner.scan_batch(usernames, force=force) # <-- Code cũ quét cả mẻ
        
        total_pass = 0
        for user in usernames:
            if _stop_event and _stop_event.is_set(): break
            res = await scanner.scan_creator(user, force=force)
            passed = len(res) if isinstance(res, list) else res.get('passed', 0)
            total_pass += passed
            _scanner_emit(f"✅ Xong @{user}: Đã bàn giao {passed} audio cho Phase 3.")
            _ai_wakeup.set() # Rung chuông báo cho AI Check worker
            
        _scanner_emit(f"🎉 Quét xong {len(usernames)} creator → Tổng {total_pass} audio tiềm năng.")

    except Exception as e:
        import traceback as tb
        _scanner_emit(f"❌ Creator Scanner lỗi: {e}")
        _scanner_emit(tb.format_exc())
    finally:
        _scanner_task = None

@app.post("/creator/scan")
async def creator_scan(body: dict):
    """Quét 1 creator."""
    global _scanner_task
    username = body.get("username", "").strip()
    force    = bool(body.get("force", False))
    if not username:
        return {"ok": False, "msg": "Thiếu username"}
    if _scanner_task and not _scanner_task.done():
        return {"ok": False, "msg": "Scanner đang chạy rồi"}
    _scanner_task = asyncio.get_event_loop().create_task(
        _run_creator_scan([username], force=force)
    )
    return {"ok": True, "msg": f"Đang quét @{username}..."}

@app.post("/creator/scan-batch")
async def creator_scan_batch(body: dict):
    """Quét nhiều creator cùng lúc."""
    global _scanner_task
    usernames = body.get("usernames", [])
    force     = bool(body.get("force", False))
    
    if usernames == ["AUTO"]:
        try:
            p = BASE_DIR / "creators_list.txt"
            if p.exists():
                usernames = [l.strip().lstrip('@') for l in p.read_text(encoding='utf-8').splitlines() if l.strip() and not l.startswith('#')]
            else:
                usernames = []
        except Exception as e:
            return {"ok": False, "msg": f"Lỗi đọc creators_list.txt: {e}"}
            
    if not usernames:
        return {"ok": False, "msg": "Thiếu usernames hoặc danh sách trống"}
        
    if _scanner_task and not _scanner_task.done():
        return {"ok": False, "msg": "Scanner đang chạy rồi"}
        
    _scanner_task = asyncio.get_event_loop().create_task(
        _run_creator_scan(usernames, force=force)
    )
    return {"ok": True, "msg": f"Đang quét {len(usernames)} creator..."}

@app.get("/creator/status")
async def creator_scan_status():
    """Trạng thái scanner + log gần nhất."""
    return {
        "running": _scanner_task is not None and not _scanner_task.done(),
        "log":     _scanner_log[-30:],
    }

@app.post("/creator/import-seen-excel")
async def creator_import_seen_excel(body: dict):
    """
    Đọc file Excel (cột A=link, cột B=username), trích xuất username
    và nạp vào danh sách ĐÃ QUÉT (seen_creators.txt) để check trùng.
    """
    file_path = body.get("file", "").strip()

    if not file_path:
        candidates = list(BASE_DIR.glob("*check trùng*.xlsx")) + \
                     list(BASE_DIR.glob("*.xlsx"))
        if not candidates:
            return {"ok": False, "msg": "Không tìm thấy file Excel 'check trùng'"}
        file_path = str(candidates[0])

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        usernames = set()
        for row in ws.iter_rows(values_only=True):
            link     = str(row[0] or "")
            username = str(row[1] or "")

            if username and username not in ("None", "Mã ID kênh ", "Người lấy "):
                clean = username.strip().lstrip("@").lower()
                if clean: usernames.add(clean)
            elif "tiktok.com/@" in link:
                try:
                    clean = link.split("tiktok.com/@")[1].split("/")[0].split("?")[0].lower()
                    if clean: usernames.add(clean)
                except: pass

        if not usernames:
            return {"ok": False, "msg": "Không trích xuất được username nào"}

        import sys
        sys.path.insert(0, str(BASE_DIR.parent / "tool_sroll_feed"))
        from creator_scanner import SEEN_CREATORS_FILE, load_seen_creators
        
        seen = load_seen_creators()
        new_users = usernames - set(seen)
        
        if new_users:
            with open(SEEN_CREATORS_FILE, "a", encoding="utf-8") as f:
                for u in new_users:
                    f.write(f"{u}\n")
                    
        return {"ok": True, "msg": f"Đã nạp {len(new_users)} kênh mới vào DB chống trùng (từ {len(usernames)} kênh trong Excel)."}
    except Exception as e:
        return {"ok": False, "msg": f"Lỗi đọc Excel: {e}"}

@app.post("/creator/import-seen-manual")
async def creator_import_seen_manual(body: dict):
    """Nạp danh sách username thủ công vào DB check trùng."""
    usernames = body.get("usernames", [])
    if not usernames:
        return {"ok": False, "msg": "Thiếu danh sách username"}
        
    try:
        import sys
        sys.path.insert(0, str(BASE_DIR.parent / "tool_sroll_feed"))
        from creator_scanner import SEEN_CREATORS_FILE, load_seen_creators
        
        seen = load_seen_creators()
        clean_users = {u.strip().lstrip("@").lower() for u in usernames if u.strip()}
        new_users = clean_users - set(seen)
        
        if new_users:
            with open(SEEN_CREATORS_FILE, "a", encoding="utf-8") as f:
                for u in new_users:
                    f.write(f"{u}\n")
                    
        return {"ok": True, "msg": f"Đã nạp {len(new_users)} kênh vào DB chống trùng."}
    except Exception as e:
        return {"ok": False, "msg": str(e)}

@app.get("/creator/seen")
async def creator_seen():
    """Danh sách creator đã quét (seen_creators.txt)."""
    try:
        import sys
        sys.path.insert(0, str(BASE_DIR.parent / "tool_sroll_feed"))
        from creator_scanner import load_seen_creators, SEEN_CREATORS_FILE
        seen = sorted(load_seen_creators())
        return {"seen": seen, "total": len(seen),
                "file": str(SEEN_CREATORS_FILE)}
    except Exception as e:
        return {"error": str(e)}

@app.delete("/creator/seen/{username}")
async def creator_unseen(username: str):
    """Xoá 1 creator khỏi seen list (để quét lại)."""
    try:
        import sys
        sys.path.insert(0, str(BASE_DIR.parent / "tool_sroll_feed"))
        from creator_scanner import SEEN_CREATORS_FILE
        clean = username.strip().lstrip("@").lower()
        if SEEN_CREATORS_FILE.exists():
            lines = [l for l in SEEN_CREATORS_FILE.read_text(encoding="utf-8").splitlines()
                     if l.strip().lstrip("@").lower() != clean]
            SEEN_CREATORS_FILE.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return {"ok": True, "msg": f"Đã xoá @{clean} khỏi seen list"}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


@app.post("/creator/import-excel")
async def creator_import_excel(body: dict):
    """
    Đọc file Excel (cột A=link, cột B=username), trích xuất username
    rồi đẩy vào hàng đợi quét.

    Body: { "file": "path/to/file.xlsx", "force": false, "skip_seen": true }
    """
    global _scanner_task
    file_path = body.get("file", "").strip()
    force     = bool(body.get("force", False))
    skip_seen = bool(body.get("skip_seen", True))  # Mặc định bỏ qua creator đã quét

    if not file_path:
        # Tìm file mặc định trong thư mục tool_nhac
        candidates = list(BASE_DIR.glob("*check trùng*.xlsx")) + \
                     list(BASE_DIR.glob("*creator*.xlsx")) + \
                     list(BASE_DIR.glob("*.xlsx"))
        if not candidates:
            return {"ok": False, "msg": "Không tìm thấy file Excel"}
        file_path = str(candidates[0])

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        usernames = []
        for row in ws.iter_rows(values_only=True):
            # Cột A: link TikTok, Cột B: username
            link     = str(row[0] or "")
            username = str(row[1] or "")

            # Ưu tiên dùng username từ cột B
            if username and username not in ("None", "Mã ID kênh ", "Người lấy "):
                clean = username.strip().lstrip("@").lower()
                if clean and clean not in usernames:
                    usernames.append(clean)
            # Fallback: parse từ link TikTok ở cột A
            elif "tiktok.com/@" in link:
                try:
                    clean = link.split("tiktok.com/@")[1].split("/")[0].split("?")[0].lower()
                    if clean and clean not in usernames:
                        usernames.append(clean)
                except Exception:
                    pass

        if not usernames:
            return {"ok": False, "msg": "Không trích xuất được username nào từ file"}

        # Lọc bỏ creator đã quét (nếu skip_seen=True)
        if skip_seen and not force:
            import sys
            sys.path.insert(0, str(BASE_DIR.parent / "tool_sroll_feed"))
            from creator_scanner import load_seen_creators
            seen = load_seen_creators()
            before = len(usernames)
            usernames = [u for u in usernames if u not in seen]
            skipped = before - len(usernames)
            _emit_log(f"📄 Excel: {before} creators → bỏ qua {skipped} đã quét → còn {len(usernames)}", "fyp")

        if not usernames:
            return {"ok": True, "msg": "Tất cả creator trong file đã được quét rồi!", "total": 0}

        if _scanner_task and not _scanner_task.done():
            return {"ok": False, "msg": "Scanner đang chạy rồi, chờ xong hoặc dừng trước"}

        _scanner_task = asyncio.get_event_loop().create_task(
            _run_creator_scan(usernames, force=force)
        )
        return {
            "ok":       True,
            "msg":      f"Đang quét {len(usernames)} creator từ {Path(file_path).name}",
            "file":     str(file_path),
            "total":    len(usernames),
            "usernames": usernames[:10],  # Preview 10 đầu
        }
    except Exception as e:
        return {"ok": False, "msg": f"Lỗi đọc Excel: {e}"}


@app.post("/creator/import-background")
async def creator_import_background(body: dict):
    """
    Chỉ dùng data Excel "làm nền":
    Thẳng vào creators_list.txt (để MarketExpander nuôi acc)
    Và seen_creators.txt (để không bị đem ra quét lại mất thời gian).
    """
    file_path = body.get("file", "").strip()
    if not file_path:
        candidates = list(BASE_DIR.glob("*check trùng*.xlsx")) + \
                     list(BASE_DIR.glob("*creator*.xlsx")) + \
                     list(BASE_DIR.glob("*.xlsx"))
        if not candidates:
            return {"ok": False, "msg": "Không tìm thấy file Excel"}
        file_path = str(candidates[0])

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        usernames = []
        for row in ws.iter_rows(values_only=True):
            link     = str(row[0] or "")
            username = str(row[1] or "")
            if username and username not in ("None", "Mã ID kênh ", "Người lấy "):
                clean = username.strip().lstrip("@").lower()
                if clean and clean not in usernames:
                    usernames.append(clean)
            elif "tiktok.com/@" in link:
                try:
                    clean = link.split("tiktok.com/@")[1].split("/")[0].split("?")[0].lower()
                    if clean and clean not in usernames:
                        usernames.append(clean)
                except Exception:
                    pass

        if not usernames:
            return {"ok": False, "msg": "Không có username nào trong file"}

        import sys
        sys.path.insert(0, str(BASE_DIR.parent / "tool_sroll_feed"))
        from creator_scanner import mark_creator_seen, save_to_shared_creators
        
        added = 0
        for u in usernames:
            save_to_shared_creators(u)
            mark_creator_seen(u)
            added += 1

        return {
            "ok": True, 
            "msg": f"Đã nạp {added} creator làm nền (không quét) từ {Path(file_path).name}"
        }
    except Exception as e:
        return {"ok": False, "msg": f"Lỗi: {e}"}

# ── WebSocket ──────────────────────────────────────────────────────────────────
@app.websocket("/ws")
async def ws_endpoint(ws: WebSocket):
    await ws.accept()
    _connected.add(ws)
    # Gửi snapshot đầy đủ cho client vừa kết nối
    await ws.send_text(json.dumps({
        "type":             "snapshot",
        "status":           _status,
        "nhac_status":      "running" if (_nhac_task and not _nhac_task.done()) else "stopped",
        "fyp_status":       _fyp_status_str(),
        "expander_status":  "running" if (_expander_task and not _expander_task.done()) else "stopped",
        "target":           _target,
        "results":          _results[:300],
        "nhac_logs":        _nhac_log[-100:],
        "fyp_logs":         _fyp_log[-100:],
        "miner_logs":       _miner_log[-100:],
        "follow_logs":      _follow_log[-50:],
        "expander_logs":    _expander_log[-50:],
        "stats":            _get_stats(),
        "creators_count":   _creators_count(),
    }, ensure_ascii=False))
    try:
        while True:
            await ws.receive_text()
    except WebSocketDisconnect:
        _connected.discard(ws)

# ── Khởi động ứng dụng ──────────────────────────────────────────────────────────
# Lifespan được định nghĩa ở trên (line 389)

# ── Static files ───────────────────────────────────────────────────────────────
static_dir = BASE_DIR / "static"
static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

@app.get("/")
async def index():
    return FileResponse(str(static_dir / "index.html"))

# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="warning")
